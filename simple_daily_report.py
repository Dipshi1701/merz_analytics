"""Standalone daily Merz report for cron — API only, no database required.

Produces the same Excel layout as dashboard Download Excel:
Row Labels | Count of User question | Labels | Product | Date | Time
(Session → Question → Answer → Rating)
"""

from __future__ import annotations

import argparse
import hashlib
import hmac
import json
import logging
import os
import smtplib
import ssl
import sys
import time
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Any
from urllib.parse import quote
from zoneinfo import ZoneInfo

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

load_dotenv(ROOT / ".env", override=True)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

os.environ.setdefault("NO_PROXY", "*")
os.environ["HTTP_PROXY"] = ""
os.environ["HTTPS_PROXY"] = ""

PRODUCT_CONTENT_IDS = {
    30: "Belotero",
    31: "DeScribe",
    32: "Radiesse",
    33: "Ultherapy",
    34: "Xeomin",
}
PRODUCT_NAMES = list(PRODUCT_CONTENT_IDS.values())
_PRODUCT_LOOKUP = {p.lower(): p for p in PRODUCT_NAMES}

COLUMNS = ["Row Labels", "Count of User question", "Labels", "Product", "Date", "Time"]
COLUMN_KEYS = {
    "Row Labels": "row_label",
    "Count of User question": "count",
    "Labels": "label_type",
    "Product": "product",
    "Date": "date",
    "Time": "time",
}
COLUMN_WIDTHS = {
    "Row Labels": 100,
    "Count of User question": 15,
    "Labels": 15,
    "Product": 14,
    "Date": 12,
    "Time": 10,
}
SESSION_FILL = PatternFill(start_color="D3E3F3", end_color="D3E3F3", fill_type="solid")


@dataclass(frozen=True)
class Settings:
    tz: ZoneInfo
    report_dir: Path
    auth_url: str
    api_key: str
    api_secret: str
    signature_key: str
    env: str
    sources: list[str] | None
    smtp_host: str
    smtp_port: int
    email_user: str
    email_password: str
    email_to: list[str]
    editor_key: str
    editor_secret: str
    editor_personal: str
    editor_base_url: str

    @classmethod
    def load(cls) -> Settings:
        required = ("MERZ_INBENTA_API_KEY", "MERZ_INBENTA_API_SECRET", "MERZ_INBENTA_SIGNATURE_KEY")
        missing = [k for k in required if not os.getenv(k)]
        if missing:
            raise SystemExit(f"Missing in .env: {', '.join(missing)}")

        raw_sources = os.getenv("MERZ_SOURCES", "").strip()
        sources = [s.strip() for s in raw_sources.split(",") if s.strip()] or None
        email_user = os.getenv("EMAIL_USER", "")
        raw_to = os.getenv("EMAIL_TO", email_user)
        recipients = [a.strip() for a in raw_to.replace(";", ",").split(",") if a.strip()]
        if not recipients:
            recipients = [email_user] if email_user else []
        return cls(
            tz=ZoneInfo(os.getenv("TIMEZONE", "America/New_York")),
            report_dir=Path(os.getenv("MERZ_REPORT_DIR", str(ROOT / "data" / "reports"))),
            auth_url=os.getenv("MERZ_INBENTA_AUTH_URL", "https://api.inbenta.io/v1/auth"),
            api_key=os.environ["MERZ_INBENTA_API_KEY"],
            api_secret=os.environ["MERZ_INBENTA_API_SECRET"],
            signature_key=os.environ["MERZ_INBENTA_SIGNATURE_KEY"],
            env=(os.getenv("MERZ_INBENTA_ENV") or "production").strip().lower(),
            sources=sources,
            smtp_host=os.getenv("SMTP_HOST", "mail.alphanumeric.ai"),
            smtp_port=int(os.getenv("SMTP_PORT", "465")),
            email_user=email_user,
            email_password=os.getenv("EMAIL_PASSWORD", "").strip().strip('"').strip("'"),
            email_to=recipients,
            editor_key=os.getenv("MERZ_INBENTA_EDITOR_API_KEY", ""),
            editor_secret=os.getenv("MERZ_INBENTA_EDITOR_API_SECRET", ""),
            editor_personal=os.getenv("MERZ_INBENTA_EDITOR_PERSONAL_SECRET_KEY", ""),
            editor_base_url=os.getenv(
                "MERZ_INBENTA_EDITOR_BASE_URL",
                "https://chatbot-api-us.inbenta.io/editor/v1",
            ).rstrip("/"),
        )


class InbentaClient:
    def __init__(self, cfg: Settings):
        self.cfg = cfg
        self._token: str | None = None
        self._token_expires = datetime.min
        self._base_url: str | None = None
        self._last_call = 0.0
        self._editor_token: str | None = None
        self._title_cache: dict[int, str] = {}

    def _auth(self) -> None:
        if self._token and datetime.now() < self._token_expires:
            return
        r = requests.post(
            self.cfg.auth_url,
            headers={"x-inbenta-key": self.cfg.api_key, "Content-Type": "application/json"},
            json={"secret": self.cfg.api_secret},
            timeout=30,
        )
        r.raise_for_status()
        data = r.json()
        self._token = data["accessToken"]
        self._token_expires = datetime.now() + timedelta(seconds=data.get("expiration", 1200) - 120)
        self._base_url = data["apis"]["reporting"].rstrip("/")

    def _sign(self, method: str, path: str, params: dict[str, str] | None) -> dict[str, str]:
        ts = int(datetime.now().timestamp())
        parts = [method.upper()]
        clean = path.split("/prod/")[-1].lstrip("/") if "/prod/" in path else path.lstrip("/")
        if clean:
            parts.append(quote(clean, safe=""))
        if params:
            qs = "&".join(f"{k}={json.dumps(v, ensure_ascii=False)}" for k, v in sorted(params.items()))
            if qs:
                parts.append(quote(qs, safe="", encoding="utf-8"))
        parts += [str(ts), "v1"]
        sig = hmac.new(
            self.cfg.signature_key.encode(),
            "&".join(parts).encode(),
            hashlib.sha256,
        ).hexdigest()
        return {
            "x-inbenta-signature": sig,
            "x-inbenta-signature-version": "v1",
            "x-inbenta-timestamp": str(ts),
        }

    def _get(self, path: str, report_date: date) -> list[dict[str, Any]]:
        self._auth()
        elapsed = time.time() - self._last_call
        if elapsed < 0.075:
            time.sleep(0.075 - elapsed)
        params = {
            "date_from": report_date.isoformat(),
            "date_to": report_date.isoformat(),
            "env": self.cfg.env,
        }
        if self.cfg.sources:
            params["source"] = json.dumps(self.cfg.sources)
        headers = {
            "Authorization": f"Bearer {self._token}",
            "Content-Type": "application/json",
            "x-inbenta-key": self.cfg.api_key,
            **self._sign("GET", path, params),
        }
        r = requests.get(f"{self._base_url}{path}", headers=headers, params=params, timeout=60)
        r.raise_for_status()
        self._last_call = time.time()
        return r.json().get("results", [])

    def questions(self, report_date: date) -> list[dict[str, Any]]:
        return self._get("/v1/events/user_questions", report_date)

    def session_details(self, report_date: date) -> list[dict[str, Any]]:
        return self._get("/v1/aggregates/session_details", report_date)

    def sessions(self, report_date: date) -> list[dict[str, Any]]:
        return self._get("/v1/events/sessions", report_date)

    def clicks(self, report_date: date) -> list[dict[str, Any]]:
        return self._get("/v1/events/clicks", report_date)

    def ratings(self, report_date: date) -> list[dict[str, Any]]:
        return self._get("/v1/events/ratings", report_date)

    def _editor_auth(self) -> bool:
        if not (self.cfg.editor_key and self.cfg.editor_secret and self.cfg.editor_personal):
            return False
        if self._editor_token:
            return True
        url = (
            f"{self.cfg.auth_url}?secret={self.cfg.editor_secret}"
            f"&user_personal_secret={self.cfg.editor_personal}"
        )
        r = requests.post(
            url,
            headers={"x-inbenta-key": self.cfg.editor_key, "Content-Type": "application/json"},
            json={
                "key": self.cfg.editor_key,
                "secret": self.cfg.editor_secret,
                "user_personal_secret": self.cfg.editor_personal,
            },
            timeout=30,
        )
        r.raise_for_status()
        self._editor_token = r.json().get("accessToken")
        return bool(self._editor_token)

    def content_title(self, content_id: int) -> str:
        if content_id in self._title_cache:
            return self._title_cache[content_id]
        title = f"Content {content_id}"
        if self._editor_auth():
            try:
                elapsed = time.time() - self._last_call
                if elapsed < 0.6:
                    time.sleep(0.6 - elapsed)
                r = requests.get(
                    f"{self.cfg.editor_base_url}/contents/{content_id}",
                    headers={
                        "Authorization": f"Bearer {self._editor_token}",
                        "x-inbenta-key": self.cfg.editor_key,
                        "Content-Type": "application/json",
                    },
                    timeout=30,
                )
                self._last_call = time.time()
                if r.status_code == 200:
                    title = (r.json().get("data") or {}).get("title") or title
            except Exception as exc:
                log.warning("Editor title fetch failed for %s: %s", content_id, exc)
        self._title_cache[content_id] = title
        return title


def parse_dt(raw: str, tz: ZoneInfo) -> datetime | None:
    if not raw:
        return None
    try:
        text = raw.strip().replace("Z", "+00:00")
        dt = datetime.fromisoformat(text)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        return dt.astimezone(tz)
    except ValueError:
        return None


def in_day(raw: str, report_date: date, tz: ZoneInfo) -> bool:
    dt = parse_dt(raw, tz)
    if not dt:
        return False
    start = datetime(report_date.year, report_date.month, report_date.day, tzinfo=tz)
    return start <= dt < start + timedelta(days=1)


def yesterday(tz: ZoneInfo) -> date:
    return datetime.now(tz).date() - timedelta(days=1)


def _normalize_product(value: str) -> str:
    cleaned = (value or "").strip()
    if not cleaned:
        return ""
    exact = _PRODUCT_LOOKUP.get(cleaned.lower())
    if exact:
        return exact
    for name in PRODUCT_NAMES:
        if cleaned.lower().startswith(name.lower()):
            return name
    return ""


def product_from_matchings(matchings: list[dict], titles: dict[int, str]) -> str:
    for m in matchings or []:
        cid = m.get("id_content")
        if cid in PRODUCT_CONTENT_IDS:
            return PRODUCT_CONTENT_IDS[cid]
        title = titles.get(cid) or ""
        product = _normalize_product(title)
        if product:
            return product
    return ""


def _parse_variable(value_str: str, name: str) -> str:
    if not value_str:
        return ""
    try:
        data = json.loads(value_str) if isinstance(value_str, str) else value_str
    except (json.JSONDecodeError, TypeError):
        return ""
    if not isinstance(data, dict):
        return ""
    for var in data.get("variables", []) or []:
        if not isinstance(var, dict):
            continue
        if str(var.get("name", "")).lower() == name.lower():
            return str(var.get("value", "")).strip()
    return ""


def map_session_products(session_events: list[dict]) -> dict[str, str]:
    """VARIABLES + SEARCH → log_id/event_id → product."""
    by_session: dict[str, list[dict]] = defaultdict(list)
    for ev in session_events:
        sid = ev.get("session_id") or ""
        if sid:
            by_session[sid].append(ev)

    out: dict[str, str] = {}
    for sid, events in by_session.items():
        events.sort(key=lambda e: (e.get("date") or "", e.get("event_id") or ""))
        current = ""
        for ev in events:
            key = (ev.get("key") or "").upper()
            if key in ("VARIABLES", "ACTION_DATA_FIELD"):
                parsed = _normalize_product(_parse_variable(ev.get("value") or "", "product"))
                if parsed:
                    current = parsed
            elif key == "SEARCH" and current:
                if ev.get("log_id"):
                    out[ev["log_id"]] = current
                if ev.get("event_id"):
                    out[ev["event_id"]] = current
                qtext = (ev.get("user_question") or "").strip().lower()
                if qtext:
                    out[f"text::{sid}::{qtext}"] = current
    return out


def group_duplicates(questions: list[dict]) -> list[tuple[str, int, list[dict]]]:
    grouped: list[tuple[str, int, list[dict]]] = []
    current_text, current_count, current_qs = None, 0, []
    for q in questions:
        text = q["user_question"]
        if text == current_text:
            current_count += 1
            current_qs.append(q)
        else:
            if current_text is not None:
                grouped.append((current_text, current_count, current_qs))
            current_text, current_count, current_qs = text, 1, [q]
    if current_text is not None:
        grouped.append((current_text, current_count, current_qs))
    return grouped


def format_answers(client: InbentaClient, questions: list[dict], clicks_by_log: dict[str, set[int]]) -> str:
    lines = []
    seen: set[int] = set()
    event_ids = [q["event_id"] for q in questions]
    for q in questions:
        matchings = sorted(
            q.get("matchings") or [],
            key=lambda m: -(m.get("weight") or 0),
        )
        for m in matchings:
            cid = m.get("id_content")
            if cid is None or cid in seen:
                continue
            seen.add(cid)
            title = client.content_title(cid)
            clicked = any(cid in clicks_by_log.get(eid, set()) for eid in event_ids)
            external = "Yes" if m.get("external") else "No"
            lines.append(f"- {title} (ID: {cid}, External: {external}, Clicks: {1 if clicked else 0})")
    return "\n".join(lines) if lines else "0"


def format_rating(questions: list[dict], ratings_by_log: dict[str, dict]) -> str:
    for q in questions:
        rating = ratings_by_log.get(q["event_id"]) or ratings_by_log.get(q.get("log_id") or "")
        if rating and rating.get("rating"):
            comment = rating.get("comment") or ""
            if comment and comment != "-":
                return f"Rating: {rating['rating']} - {comment}"
            return f"Rating: {rating['rating']}"
    return ""


def build_report_rows(cfg: Settings, report_date: date) -> list[dict[str, Any]]:
    client = InbentaClient(cfg)
    log.info("Fetching API data for %s (env=%s) — no DB", report_date, cfg.env)

    raw_questions = [
        q for q in client.questions(report_date)
        if q.get("event_id") and in_day(q.get("date", ""), report_date, cfg.tz)
        and (q.get("user_question") or "").strip()
    ]
    details = [
        s for s in client.session_details(report_date)
        if in_day(s.get("date", ""), report_date, cfg.tz) or not s.get("date")
    ]
    session_events = client.sessions(report_date)
    clicks = client.clicks(report_date)
    ratings = client.ratings(report_date)

    questions_by_id = {}
    for q in raw_questions:
        eid = q["event_id"]
        questions_by_id[eid] = {
            "event_id": eid,
            "user_question": (q.get("user_question") or "").strip(),
            "date": q.get("date", ""),
            "log_id": q.get("log_id") or "",
            "matchings": q.get("matchings") or [],
        }

    # Pre-resolve titles used for product + answers
    title_map: dict[int, str] = {}
    content_ids = {
        m.get("id_content")
        for q in questions_by_id.values()
        for m in q["matchings"]
        if m.get("id_content") is not None
    }
    for cid in content_ids:
        title_map[cid] = client.content_title(cid)

    product_from_session = map_session_products(session_events)

    clicks_by_log: dict[str, set[int]] = defaultdict(set)
    for c in clicks:
        lid = c.get("log_id") or ""
        cid = c.get("id_content")
        if lid and cid is not None:
            clicks_by_log[lid].add(cid)

    ratings_by_log: dict[str, dict] = {}
    for r in ratings:
        lid = r.get("log_id") or ""
        if lid and r.get("rating") is not None:
            ratings_by_log[lid] = {"rating": r.get("rating"), "comment": r.get("comment") or ""}

    # session_id -> questions
    session_map: dict[str, list[dict]] = defaultdict(list)
    linked: set[str] = set()
    session_meta: dict[str, dict] = {}

    for s in details:
        sid = s.get("session_id") or ""
        if not sid:
            continue
        log_ids = s.get("log_ids") or []
        if isinstance(log_ids, str):
            try:
                log_ids = json.loads(log_ids)
            except (json.JSONDecodeError, TypeError):
                log_ids = []
        data_keys = s.get("data_keys") or []
        if isinstance(data_keys, str):
            try:
                data_keys = json.loads(data_keys)
            except (json.JSONDecodeError, TypeError):
                data_keys = []
        escalated = (
            "Yes/Attended" if "CHAT_ATTENDED" in data_keys
            else "Yes" if "CHAT_NO_AGENTS" in data_keys else ""
        )
        session_meta[sid] = {"date": s.get("date", ""), "escalated": escalated}
        for eid in log_ids:
            linked.add(eid)
            if eid in questions_by_id:
                session_map[sid].append(questions_by_id[eid])

    for eid, q in questions_by_id.items():
        if eid not in linked:
            session_map["No Session"].append(q)
            session_meta.setdefault("No Session", {"date": q.get("date", ""), "escalated": ""})

    rows: list[dict[str, Any]] = []
    for sid, qs in session_map.items():
        if not qs:
            continue
        qs.sort(key=lambda x: x.get("date") or "")
        meta = session_meta.get(sid, {})
        dt = parse_dt(meta.get("date") or (qs[0].get("date") if qs else ""), cfg.tz)
        date_str = dt.strftime("%m/%d/%Y") if dt else report_date.strftime("%m/%d/%Y")
        time_str = dt.strftime("%H:%M:%S") if dt else ""

        rows.append({
            "row_label": sid,
            "count": len(qs),
            "label_type": "Session",
            "product": "",
            "date": date_str,
            "time": time_str,
        })

        for question_text, count, group_qs in group_duplicates(qs):
            product = ""
            for q in group_qs:
                product = product_from_matchings(q["matchings"], title_map)
                if product:
                    break
            if not product:
                for q in group_qs:
                    product = (
                        product_from_session.get(q["event_id"])
                        or product_from_session.get(q.get("log_id") or "")
                        or product_from_session.get(
                            f"text::{sid}::{(q.get('user_question') or '').strip().lower()}",
                            "",
                        )
                    )
                    if product:
                        break

            rows.append({
                "row_label": question_text,
                "count": count,
                "label_type": "Question",
                "product": product,
                "date": "",
                "time": "",
            })
            rows.append({
                "row_label": format_answers(client, group_qs, clicks_by_log),
                "count": count,
                "label_type": "Answer",
                "product": "",
                "date": "",
                "time": "",
            })
            rows.append({
                "row_label": format_rating(group_qs, ratings_by_log),
                "count": count,
                "label_type": "Rating",
                "product": "",
                "date": "",
                "time": "",
            })

    session_count = sum(1 for r in rows if r["label_type"] == "Session")
    log.info("%d sessions, %d questions", session_count, len(questions_by_id))
    return rows


def write_excel(cfg: Settings, rows: list[dict[str, Any]], report_date: date) -> Path:
    cfg.report_dir.mkdir(parents=True, exist_ok=True)
    path = cfg.report_dir / f"merz-report_{report_date.strftime('%Y-%m-%d')}_to_{report_date.strftime('%Y-%m-%d')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    for col_idx, header in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="top")

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            key = COLUMN_KEYS[col_name]
            default = 1 if col_name == "Count of User question" else ""
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, default))
        if row_data.get("label_type") == "Session":
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = SESSION_FILL

    for col_idx, header in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS.get(header, 15)
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows) + 1}"
    for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1, max_col=len(COLUMNS)):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    wb.save(path)
    log.info("Saved %s", path)
    return path


NO_DATA_MSG = "No questions recorded for this date."


def email_body(report_date: date, count: int, filename: str | None) -> tuple[str, str]:
    label = report_date.strftime("%A, %B %d, %Y")
    short = report_date.strftime("%b %d, %Y")
    if count == 0:
        note = NO_DATA_MSG
        banner = (
            f'<p style="margin:0;padding:12px;background:#fff3cd;color:#856404;'
            f'border-left:4px solid #ffc107">{NO_DATA_MSG}</p>'
        )
        attach = ""
    else:
        note = f"Sessions: {count}"
        banner = (
            f'<p style="margin:0;padding:12px;background:#d4edda;color:#155724;'
            f'border-left:4px solid #28a745"><b>{count}</b> session(s). Excel attached.</p>'
        )
        attach = (
            f"<tr><td style='padding:10px;color:#888'>Attachment</td>"
            f"<td style='padding:10px'>{filename}</td></tr>"
        )

    plain = f"Merz Daily Report\n\nDate: {label}\n{note}\n"
    html = f"""<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;background:#f4f6f9;padding:24px">
<table width="600" style="margin:0 auto;background:#fff;border-radius:8px">
<tr><td style="background:#1a3a5c;color:#fff;padding:24px"><h2 style="margin:0">Merz Chatbot</h2>
<p style="margin:4px 0 0;opacity:.8">Daily Report</p></td></tr>
<tr><td style="padding:24px"><p>Summary for <b>{short}</b></p>
<table width="100%" style="border:1px solid #e8ecf0;margin-bottom:16px">
<tr><td style="padding:10px;color:#888">Report Date</td><td style="padding:10px"><b>{label}</b></td></tr>
<tr><td style="padding:10px;color:#888">Sessions</td><td style="padding:10px"><b>{count}</b></td></tr>
{attach}</table>{banner}</td></tr></table></body></html>"""
    return plain, html


def send_mail(cfg: Settings, report_date: date, count: int, path: Path | None = None) -> None:
    if not cfg.email_user or not cfg.email_password:
        raise SystemExit("EMAIL_USER and EMAIL_PASSWORD required in .env")
    if not cfg.email_to:
        raise SystemExit("EMAIL_TO required in .env")

    plain, html = email_body(report_date, count, path.name if path else None)
    msg = MIMEMultipart("mixed")
    subject = f"Merz Daily Report - {report_date.strftime('%b %d, %Y')}"
    if count == 0:
        subject += " - No activity"
    msg["Subject"] = subject
    msg["From"] = cfg.email_user
    msg["To"] = ", ".join(cfg.email_to)

    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(plain, "plain", "utf-8"))
    alt.attach(MIMEText(html, "html", "utf-8"))
    msg.attach(alt)

    if path:
        with path.open("rb") as f:
            part = MIMEApplication(f.read(), _subtype="xlsx")
            part.add_header("Content-Disposition", "attachment", filename=path.name)
            msg.attach(part)

    with smtplib.SMTP_SSL(
        cfg.smtp_host, cfg.smtp_port, context=ssl.create_default_context(), timeout=60
    ) as s:
        s.login(cfg.email_user, cfg.email_password)
        s.sendmail(cfg.email_user, cfg.email_to, msg.as_string())
    log.info("Email sent to %s", ", ".join(cfg.email_to))


def run(report_date: date | None = None, send: bool = True) -> Path | None:
    cfg = Settings.load()
    day = report_date or yesterday(cfg.tz)
    rows = build_report_rows(cfg, day)
    session_count = sum(1 for r in rows if r.get("label_type") == "Session")

    if session_count == 0:
        log.info(NO_DATA_MSG)
        if send:
            send_mail(cfg, day, 0)
        return None

    path = write_excel(cfg, rows, day)
    if send:
        send_mail(cfg, day, session_count, path)
    return path


def main() -> None:
    p = argparse.ArgumentParser(description="Merz daily Excel report (API-only, no DB)")
    p.add_argument("--date", help="YYYY-MM-DD (default: yesterday)")
    p.add_argument("--no-email", action="store_true")
    args = p.parse_args()
    day = date.fromisoformat(args.date) if args.date else None
    result = run(day, send=not args.no_email)
    if result:
        print(result)


if __name__ == "__main__":
    main()
