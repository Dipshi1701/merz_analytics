"""Excel report generation: MySQL connection shim, session builder, and writer."""

import json
import logging
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

import pymysql
from pymysql.cursors import DictCursor
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from reporting_dashboard.config import DATABASE_URL, REPORT_DIR, get_week_key

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# MySQL shim (translates sqlite3-style API to MySQL for SessionBuilder)
# ---------------------------------------------------------------------------

def _rewrite_sql(sql: str) -> str:
    if "INSERT OR REPLACE INTO content_lookup" in sql:
        sql = sql.replace(
            "INSERT OR REPLACE INTO content_lookup (id_content, title, last_updated)",
            "INSERT INTO content_lookup (id_content, title, last_updated)",
        )
        sql = sql.replace(
            "VALUES (?, ?, CURRENT_TIMESTAMP)",
            "VALUES (%s, %s, NOW()) ON DUPLICATE KEY UPDATE title=VALUES(title), updated_at=NOW()",
        )
    for old, new in [(" key ", " `key` "), (" key\n", " `key`\n"), ("(key,", "(`key`,"),
                     (", key,", ", `key`,"), (", key)", ", `key`)")]:
        sql = sql.replace(old, new)
    return sql.replace("?", "%s")


def _parse_db_url(url: str) -> dict:
    url = url.replace("mysql+pymysql://", "")
    user_pass, host_db = url.split("@", 1)
    user, password = user_pass.split(":", 1)
    host_port, database = host_db.split("/", 1)
    host, port = (host_port.split(":") if ":" in host_port else (host_port, "3306"))
    return dict(user=user, password=password, host=host, port=int(port), database=database)


class _Cursor:
    def __init__(self, cursor):
        self._c = cursor

    def execute(self, sql, params=None):
        return self._c.execute(_rewrite_sql(sql), params or ())

    def executemany(self, sql, params):
        return self._c.executemany(_rewrite_sql(sql), params)

    def fetchone(self):
        row = self._c.fetchone()
        return tuple(row.values()) if isinstance(row, dict) else row

    def fetchall(self):
        rows = self._c.fetchall()
        if rows and isinstance(rows[0], dict):
            return [tuple(r.values()) for r in rows]
        return rows

    @property
    def rowcount(self):
        return self._c.rowcount


class MySQLConnection:
    """sqlite3.Connection drop-in backed by MySQL — used only by SessionBuilder."""

    def __init__(self):
        self._conn = pymysql.connect(charset="utf8mb4", cursorclass=DictCursor, **_parse_db_url(DATABASE_URL))

    def cursor(self):
        return _Cursor(self._conn.cursor())

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        self._conn.close()

    def __enter__(self):
        return self

    def __exit__(self, *args):
        self.close()


# ---------------------------------------------------------------------------
# Content resolver
# ---------------------------------------------------------------------------

class ContentResolver:
    def __init__(self, conn, editor_client=None):
        self.conn = conn
        self.editor_client = editor_client
        self._cache: Dict[int, str] = {}
        cursor = conn.cursor()
        cursor.execute("SELECT id_content, title FROM content_lookup")
        self._cache = {row[0]: row[1] for row in cursor.fetchall()}

    def get_title(self, id_content: int) -> str:
        if id_content in self._cache:
            return self._cache[id_content]
        cursor = self.conn.cursor()
        cursor.execute("SELECT title FROM content_lookup WHERE id_content = %s", (id_content,))
        row = cursor.fetchone()
        if row:
            self._cache[id_content] = row[0]
            return row[0]
        return f"Content {id_content}"

    def resolve_from_matchings(self, week_key: str):
        cursor = self.conn.cursor()
        cursor.execute("SELECT DISTINCT id_content FROM raw_uq_matchings WHERE week_key = %s", (week_key,))
        all_ids = [row[0] for row in cursor.fetchall()]
        ids_to_fetch = [cid for cid in all_ids if cid not in self._cache or self._cache[cid].startswith("Content ")]
        if not ids_to_fetch:
            return
        content_map = self.editor_client.get_contents_batch(ids_to_fetch) if self.editor_client else {}
        for cid in ids_to_fetch:
            title = content_map[cid].get("title", f"Content {cid}") if cid in content_map else f"Content {cid}"
            cursor.execute(
                "INSERT INTO content_lookup (id_content, title, updated_at) VALUES (%s, %s, NOW()) "
                "ON DUPLICATE KEY UPDATE title = VALUES(title), updated_at = NOW()",
                (cid, title),
            )
            self._cache[cid] = title
        self.conn.commit()


# ---------------------------------------------------------------------------
# Session builder
# ---------------------------------------------------------------------------

class SessionBuilder:
    def __init__(self, conn, editor_client=None):
        self.conn = conn
        self.content_resolver = ContentResolver(conn, editor_client=editor_client)

    def build_report_rows(self, week_key: str) -> List[Dict[str, Any]]:
        self.content_resolver.resolve_from_matchings(week_key)
        user_questions = self._get_user_questions(week_key)
        sessions = self._get_sessions(week_key)
        session_questions_map = self._link_questions(user_questions, sessions)

        rows = []
        for session in sessions:
            sid = session["session_id"]
            questions_in_session = session_questions_map.get(sid, [])
            if not questions_in_session:
                continue
            date_str, time_str = self._parse_dt(session["date"])
            user_type = self._get_user_type(sid, week_key)
            rows.append({"row_label": sid, "count": len(questions_in_session), "label_type": "Session",
                         "user_type": user_type, "date": date_str,
                         "time": time_str, "escalated": session.get("escalated", ""), "product": ""})
            for question_text, count, event_ids in self._group_duplicates(questions_in_session):
                product = self._get_product_context(event_ids)
                rows.append({"row_label": question_text, "count": count, "label_type": "Question",
                             "product": product})
                rows.append({"row_label": self._format_matchings(event_ids, week_key), "count": count,
                             "label_type": "Answer", "product": ""})
                rows.append({"row_label": self._get_rating(event_ids, week_key), "count": count,
                             "label_type": "Rating", "product": ""})
        return rows

    def _get_user_questions(self, week_key: str) -> List[Dict]:
        cursor = self.conn.cursor()
        cursor.execute(
            "SELECT r.event_id, r.user_question, r.date, COALESCE(u.product_context, '') "
            "FROM raw_user_questions r "
            "LEFT JOIN user_questions u ON u.event_id = r.event_id "
            "WHERE r.week_key = %s ORDER BY r.date ASC",
            (week_key,),
        )
        return [{"event_id": r[0], "user_question": r[1], "date": r[2], "product_context": r[3] or ""}
                for r in cursor.fetchall()]

    def _get_product_context(self, event_ids: List[str]) -> str:
        if not event_ids:
            return ""
        cursor = self.conn.cursor()
        ph = ",".join(["%s"] * len(event_ids))
        cursor.execute(
            f"SELECT product_context FROM user_questions "
            f"WHERE event_id IN ({ph}) AND product_context IS NOT NULL AND product_context != '' LIMIT 1",
            tuple(event_ids),
        )
        row = cursor.fetchone()
        return row[0] if row and row[0] else ""

    def _get_sessions(self, week_key: str) -> List[Dict]:
        cursor = self.conn.cursor()
        cursor.execute("SELECT session_id, source, date, log_ids, ids_contents, data_keys "
                       "FROM agg_session_details WHERE week_key = %s ORDER BY date, session_id", (week_key,))
        sessions = []
        for row in cursor.fetchall():
            try:
                data_keys = json.loads(row[5]) if row[5] else []
            except (json.JSONDecodeError, TypeError):
                data_keys = []
            escalated = ("Yes/Attended" if "CHAT_ATTENDED" in data_keys
                         else "Yes" if "CHAT_NO_AGENTS" in data_keys else "")
            sessions.append({"session_id": row[0], "source": row[1], "date": row[2],
                              "log_ids": row[3], "escalated": escalated})
        return sessions

    def _get_user_type(self, session_id: str, week_key: str) -> str:
        cursor = self.conn.cursor()
        cursor.execute("SELECT value FROM raw_sessions WHERE session_id = %s "
                       "AND `key` IN ('VARIABLES', 'ACTION_DATA_FIELD') AND week_key = %s LIMIT 10",
                       (session_id, week_key))
        for (value_str,) in cursor.fetchall():
            if not value_str:
                continue
            try:
                data = json.loads(value_str) if isinstance(value_str, str) else value_str
                for var in data.get("variables", []):
                    if isinstance(var, dict) and var.get("name", "").lower() == "usertype":
                        val = var.get("value", "").strip().upper()
                        if val in ("CONSUMER", "HCP"):
                            return val
            except (json.JSONDecodeError, TypeError, AttributeError):
                continue
        return ""

    def _link_questions(self, user_questions: List[Dict], sessions: List[Dict]) -> Dict[str, List]:
        session_map = defaultdict(list)
        for q in user_questions:
            for s in sessions:
                if not s.get("log_ids"):
                    continue
                try:
                    if q["event_id"] in json.loads(s["log_ids"]):
                        session_map[s["session_id"]].append(q)
                        break
                except (json.JSONDecodeError, TypeError):
                    continue
        return session_map

    def _group_duplicates(self, questions: List[Dict]) -> List[Tuple[str, int, List[str]]]:
        grouped, current_text, current_count, current_ids = [], None, 0, []
        for q in questions:
            text = q["user_question"]
            if text == current_text:
                current_count += 1
                current_ids.append(q["event_id"])
            else:
                if current_text is not None:
                    grouped.append((current_text, current_count, current_ids))
                current_text, current_count, current_ids = text, 1, [q["event_id"]]
        if current_text is not None:
            grouped.append((current_text, current_count, current_ids))
        return grouped

    def _parse_dt(self, dt_str: str) -> Tuple[str, str]:
        if not dt_str:
            return "", ""
        try:
            s = dt_str.strip().replace("Z", "+00:00")
            dt = datetime.fromisoformat(s)
            if "T" not in s:
                return dt.strftime("%m/%d/%Y"), "00:00:00"
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=ZoneInfo("UTC"))
            dt_et = dt.astimezone(ZoneInfo("America/New_York"))
            return dt_et.strftime("%m/%d/%Y"), dt_et.strftime("%H:%M:%S")
        except (ValueError, AttributeError):
            return "", ""

    def _format_matchings(self, event_ids: List[str], week_key: str) -> str:
        if not event_ids:
            return "0"
        cursor = self.conn.cursor()
        ph = ",".join(["%s"] * len(event_ids))
        cursor.execute(f"SELECT id_content, is_external FROM raw_uq_matchings "
                       f"WHERE event_id IN ({ph}) AND week_key = %s ORDER BY weight DESC",
                       (*event_ids, week_key))
        rows = cursor.fetchall()
        if not rows:
            return "0"
        lines = []
        for id_content, is_external in rows:
            title = self.content_resolver.get_title(id_content)
            cursor2 = self.conn.cursor()
            ph2 = ",".join(["%s"] * len(event_ids))
            cursor2.execute(f"SELECT COUNT(*) FROM raw_clicks WHERE log_id IN ({ph2}) AND id_content = %s AND week_key = %s",
                            (*event_ids, id_content, week_key))
            row2 = cursor2.fetchone()
            clicks = 1 if row2 and row2[0] > 0 else 0
            lines.append(f"- {title} (ID: {id_content}, External: {'Yes' if is_external else 'No'}, Clicks: {clicks})")
        return "\n".join(lines)

    def _get_rating(self, event_ids: List[str], week_key: str) -> str:
        if not event_ids:
            return ""
        cursor = self.conn.cursor()
        ph = ",".join(["%s"] * len(event_ids))
        cursor.execute(f"SELECT rating, comment FROM raw_ratings WHERE log_id IN ({ph}) AND week_key = %s LIMIT 1",
                       (*event_ids, week_key))
        row = cursor.fetchone()
        if row and row[0]:
            return f"Rating: {row[0]} - {row[1]}" if row[1] and row[1] != "-" else f"Rating: {row[0]}"
        return ""


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

class MerzExcelWriter:
    COLUMNS = ["Row Labels", "Count of User question", "Labels", "Product", "Date", "Time"]
    COLUMN_DATA_KEYS = {
        "Row Labels": "row_label", "Count of User question": "count", "Labels": "label_type",
        "Product": "product", "Date": "date", "Time": "time",
    }
    COLUMN_WIDTHS = {
        "Row Labels": 100, "Count of User question": 15, "Labels": 15,
        "Product": 14, "Date": 12, "Time": 10,
    }
    SESSION_FILL = PatternFill(start_color="D3E3F3", end_color="D3E3F3", fill_type="solid")

    def __init__(self, output_dir: str):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def write_report(self, rows: List[Dict], date_from: datetime, date_to: datetime, custom_filename: str = None) -> str:
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        for col_idx, header in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="top")
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(self.COLUMNS, 1):
                key = self.COLUMN_DATA_KEYS.get(col_name, col_name.lower())
                default = 1 if col_name == "Count of User question" else ""
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, default))
            if row_data.get("label_type") == "Session":
                for col_idx in range(1, len(self.COLUMNS) + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = self.SESSION_FILL
        for col_idx, header in enumerate(self.COLUMNS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = self.COLUMN_WIDTHS.get(header, 15)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(self.COLUMNS))}{len(rows) + 1}"
        for row in ws.iter_rows(min_row=2, max_row=len(rows) + 1, max_col=len(self.COLUMNS)):
            for cell in row:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        filename = custom_filename or f"merz_weekly_report_{date_from.strftime('%d%m%Y')}_{date_to.strftime('%d%m%Y')}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)
        logger.info(f"Excel saved: {filepath}")
        return str(filepath)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def _week_keys_for_range(start_date: date, end_date: date) -> List[str]:
    keys = set()
    current = start_date
    while current <= end_date:
        keys.add(get_week_key(datetime.combine(current, datetime.min.time())))
        current += timedelta(days=1)
    return sorted(keys)


def generate_excel(start_date: date, end_date: date) -> str:
    """Always sync from API then build Excel."""
    from reporting_dashboard.sync import sync
    from reporting_dashboard.api import create_editor_client

    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    sync(start_date, end_date)

    editor_client = create_editor_client()
    report_rows: List[Dict[str, Any]] = []
    with MySQLConnection() as conn:
        builder = SessionBuilder(conn, editor_client=editor_client)
        for week_key in _week_keys_for_range(start_date, end_date):
            report_rows.extend(builder.build_report_rows(week_key))

    writer = MerzExcelWriter(str(REPORT_DIR))
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.min.time())
    filename = f"merz-report_{start_date.strftime('%Y-%m-%d')}_to_{end_date.strftime('%Y-%m-%d')}.xlsx"
    return writer.write_report(rows=report_rows, date_from=start_dt, date_to=end_dt, custom_filename=filename)
